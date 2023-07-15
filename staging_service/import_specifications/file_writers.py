"""
Write an import specification to one or more files.

The names of the files will be the datatype suffixed by the file extension unless the writer
handles Excel or similar files that can contain multiple datatypes, in which case the
file name will be import_specification suffixed by the extension.

All the write_* functions in this module have the same function signature:

:param folder: Where the files should be written. The folder must exist.
:param types: The import specifications to write. This is a dictionary of data types as strings
    to the specifications for the data type. Each specification has two required keys:
    * `order_and_display`: this is a list of lists. Each inner list has two elements:
        * The parameter ID of a parameter. This is typically the `id` field from the
            KBase app `spec.json` file.
        * The display name of the parameter. This is typically the `ui-name` field from the
            KBase app `display.yaml` file.
        The order of the inner lists in the outer list defines the order of the columns
        in the resulting import specification files.
    * `data`: this is a list of str->str or number dicts. The keys of the dicts are the
        parameter IDs as described above, while the values are the values of the parameters.
        Each dict must have exactly the same keys as the `order_and_display` structure. Each
        entry in the list corresponds to a row in the resulting import specification,
        and the order of the list defines the order of the rows.
    Leave the `data` list empty to write an empty template.
:returns: A mapping of the data types to the files to which they were written.
"""
# note that we can't use an f string here to interpolate the variables below, e.g.
# order_and_display, etc.

import collections
import csv
import numbers
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# this version is synonymous to the versions in individual_parsers.py. However, this module
# should only ever write the most recent format for import specifictions, while the parsers
# may need to also be able to parse earlier versions.
_VERSION = 1

# these are the same as in individual_parsers.py. They might change from version to version so
# have a separate copy here.
_DATA_TYPE = "Data type:"
_VERSION_STR = "Version:"
_COLUMN_STR = "Columns:"
_HEADER_SEP = ";"

_IMPORT_SPEC_FILE_NAME = "import_specification"

_ORDER_AND_DISPLAY = "order_and_display"
_DATA = "data"
_EXT_CSV = "csv"
_EXT_TSV = "tsv"
_EXT_EXCEL = "xlsx"
_SEP_CSV = ","
_SEP_TSV = "\t"


def _check_import_specification(types: dict[str, dict[str, list[Any]]]):
    """
    Check the structure of an import specification data structure. If the input is empty the
    result is a noop.

    :param types: The import specifications to check. This is a dictionary of data types as strings
        to the specifications for the data type. Each specification has two required keys:
        * order_and_display: this is a list of lists. Each inner list has two elements:
            * The parameter ID of a parameter. This is typically the `id` field from the
                KBase app `spec.json` file.
            * The display name of the parameter. This is typically the `ui-name` field from the
                KBase app `display.yaml` file.
            The order of the inner lists in the outer list defines the order of the columns
            in the resulting import specification files.
        * data: this is a list of str->str or number dicts. The keys of the dicts are the
            parameter IDs as described above, while the values are the values of the parameters.
            Each dict must have exactly the same keys as the order_and_display structure.
            Each entry in the list corresponds to a row in the resulting import specification,
            and the order of the list defines the order of the rows.
        Leave the data list empty to write an empty template.
    """
    if not types:
        raise ImportSpecWriteException("At least one data type must be specified")
    for datatype in types:
        # replace this with jsonschema? don't worry about it for now
        _check_string(datatype, "A data type")
        spec = types[datatype]
        if not isinstance(spec, dict):
            raise ImportSpecWriteException(
                f"The value for data type {datatype} must be a mapping"
            )
        if _ORDER_AND_DISPLAY not in spec:
            raise ImportSpecWriteException(
                f"Data type {datatype} missing {_ORDER_AND_DISPLAY} key"
            )
        _check_is_sequence(
            spec[_ORDER_AND_DISPLAY], f"Data type {datatype} {_ORDER_AND_DISPLAY} value"
        )
        if len(spec[_ORDER_AND_DISPLAY]) == 0:
            raise ImportSpecWriteException(
                f"At least one entry is required for {_ORDER_AND_DISPLAY} for type {datatype}"
            )
        if _DATA not in spec:
            raise ImportSpecWriteException(f"Data type {datatype} missing {_DATA} key")
        _check_is_sequence(spec[_DATA], f"Data type {datatype} {_DATA} value")

        param_ids = set()
        for i, id_display in enumerate(spec[_ORDER_AND_DISPLAY]):
            err = (
                f"Invalid {_ORDER_AND_DISPLAY} entry for datatype {datatype} "
                + f"at index {i} "
            )
            _check_is_sequence(id_display, err + "- the entry")
            if len(id_display) != 2:
                raise ImportSpecWriteException(err + "- expected 2 item list")
            pid = id_display[0]
            _check_string(pid, err + "- parameter ID")
            _check_string(id_display[1], err + "- parameter display name")
            param_ids.add(pid)
        for i, datarow in enumerate(spec[_DATA]):
            err = f"Data type {datatype} {_DATA} row {i}"
            if not isinstance(datarow, dict):
                raise ImportSpecWriteException(err + " is not a mapping")
            if datarow.keys() != param_ids:
                raise ImportSpecWriteException(
                    err + f" does not have the same keys as {_ORDER_AND_DISPLAY}"
                )
            for pid, v in datarow.items():
                if (
                    v is not None
                    and not isinstance(v, numbers.Number)
                    and not isinstance(v, str)
                ):
                    raise ImportSpecWriteException(
                        err
                        + f"'s value for parameter {pid} is not a number or a string"
                    )


def _check_string(tocheck: Any, errprefix: str):
    if not isinstance(tocheck, str) or not tocheck.strip():
        raise ImportSpecWriteException(
            errprefix + " cannot be a non-string or a whitespace only string"
        )


def _check_is_sequence(tocheck: Any, errprefix: str):
    if not (
        isinstance(tocheck, collections.abc.Sequence) and not isinstance(tocheck, str)
    ):
        raise ImportSpecWriteException(errprefix + " is not a list")


def write_csv(folder: Path, types: dict[str, dict[str, list[Any]]]) -> dict[str, Path]:
    """
    Writes import specifications to 1 or more csv files. All the writers in this module
    have the same function signatures; see the module level documentation.
    """
    return _write_xsv(folder, types, _EXT_CSV, _SEP_CSV)


def write_tsv(folder: Path, types: dict[str, dict[str, list[Any]]]) -> dict[str, Path]:
    """
    Writes import specifications to 1 or more tsv files. All the writers in this module
    have the same function signatures; see the module level documentation.
    """
    return _write_xsv(folder, types, _EXT_TSV, _SEP_TSV)


def _write_xsv(
    folder: Path, types: dict[str, dict[str, list[Any]]], ext: str, sep: str
):
    _check_write_args(folder, types)
    res = {}
    for datatype in types:
        filename = datatype + "." + ext
        dt = types[datatype]
        cols = len(dt[_ORDER_AND_DISPLAY])
        with open(folder / filename, "w", newline="", encoding="utf-8") as f:
            csvw = csv.writer(f, delimiter=sep)  # handle sep escaping
            csvw.writerow(
                [
                    f"{_DATA_TYPE} {datatype}{_HEADER_SEP} "
                    + f"{_COLUMN_STR} {cols}{_HEADER_SEP} {_VERSION_STR} {_VERSION}"
                ]
            )
            pids = [i[0] for i in dt[_ORDER_AND_DISPLAY]]
            csvw.writerow(pids)
            csvw.writerow([i[1] for i in dt[_ORDER_AND_DISPLAY]])
            for row in dt[_DATA]:
                csvw.writerow([row[pid] for pid in pids])
        res[datatype] = filename
    return res


def _check_write_args(folder: Path, types: dict[str, dict[str, list[Any]]]):
    if not folder:
        # this is a programming error, not a user input error, so not using the custom
        # exception here
        raise ValueError("The folder cannot be null")
    if not isinstance(types, dict):
        raise ImportSpecWriteException("The types value must be a mapping")
    _check_import_specification(types)


def write_excel(
    folder: Path, types: dict[str, dict[str, list[Any]]]
) -> dict[str, Path]:
    """
    Writes import specifications to an Excel files. All the writers in this module
    have the same function signatures; see the module level documentation.
    """
    _check_write_args(folder, types)
    res = {}
    filename = _IMPORT_SPEC_FILE_NAME + "." + _EXT_EXCEL
    wb = Workbook()
    for datatype in types:
        dt = types[datatype]
        cols = len(dt[_ORDER_AND_DISPLAY])
        wb.create_sheet(datatype)
        sheet = wb[datatype]
        _write_excel_row(sheet, 3, [i[1] for i in dt[_ORDER_AND_DISPLAY]])
        pids = [i[0] for i in dt[_ORDER_AND_DISPLAY]]
        for xlrow, row in enumerate(dt[_DATA], start=4):
            # order by parameter id in the order_and_display list
            _write_excel_row(sheet, xlrow, [row[pid] for pid in pids])
        _expand_excel_columns_to_max_width(sheet)
        # Add the hidden data *after* expanding the columns
        sheet["A1"] = (
            f"{_DATA_TYPE} {datatype}{_HEADER_SEP} "
            + f"{_COLUMN_STR} {cols}{_HEADER_SEP} {_VERSION_STR} {_VERSION}"
        )
        _write_excel_row(sheet, 2, pids)
        sheet.row_dimensions[1].hidden = True
        sheet.row_dimensions[2].hidden = True
        res[datatype] = filename
    # trash the automatically created sheet
    wb.remove(wb[wb.sheetnames[0]])
    wb.save(folder / filename)
    return res


def _write_excel_row(sheet: Worksheet, row: int, contents: list[Any]):
    # https://stackoverflow.com/a/33921552/643675
    for col, val in enumerate(contents, start=1):
        sheet.cell(row=row, column=col).value = val


def _expand_excel_columns_to_max_width(sheet: Worksheet):
    # https://stackoverflow.com/a/40935194/643675
    for column_cells in sheet.columns:
        length = max(len(_as_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = length


def _as_text(value):
    return "" if value is None else str(value)


class ImportSpecWriteException(Exception):
    """
    An exception thrown when writing an import specification fails.
    """
